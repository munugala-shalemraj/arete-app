"""
Arete Research Study — Statistical Analysis Script
====================================================
Newcastle University MSc Computer Science (CSC8639)
Researcher : Shalem Raj Munugala
Supervisor : Dr Sara Johansson Fernstad

Analyses:
  H1 — IMI motivation scores vs published norms (descriptive statistics)
  H2 — Pre/post knowledge test learning gain (paired t-test)
  H3 — SDT (IMI subscales) vs engagement metrics (Spearman, supplementary)
  +  — Pseudonymised coding sheet for manual thematic analysis
  +  — SUS usability vs benchmark (68), skill mastery, quiz difficulty

Outputs:
  • Console report with all results
  • arete_analysis_results_<timestamp>.xlsx  — results tables
  • arete_analysis_charts_<timestamp>.pdf    — all figures

Usage:
    pip install requests pandas scipy matplotlib seaborn openpyxl
    python research_analysis.py
"""

import os, sys, getpass, warnings
from datetime import datetime
warnings.filterwarnings("ignore")

# Compatibility shim: matplotlib 3.9+ removed register_cmap; patch it back for seaborn
import matplotlib.cm as _mpl_cm
if not hasattr(_mpl_cm, "register_cmap"):
    _mpl_cm.register_cmap = lambda *a, **kw: None  # no-op

# ── Dependency check ──────────────────────────────────────────────────────────
try:
    import requests
    import pandas as pd
    import numpy as np
    import matplotlib.pyplot as plt
    import matplotlib.gridspec as gridspec
    from matplotlib.backends.backend_pdf import PdfPages
    import seaborn as sns
    from scipy import stats
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Run: pip install requests pandas scipy matplotlib seaborn openpyxl")
    sys.exit(1)

# ── Configuration ─────────────────────────────────────────────────────────────

SUPABASE_URL = "https://iscrbkcqzkqludqoobnj.supabase.co"
IMI_NORM     = 4.0     # Published IMI norm for interest/enjoyment (Ryan & Deci)
SUS_BENCH    = 68.0    # Industry SUS benchmark (Bangor et al., 2008)
ALPHA        = 0.05    # Significance level

# ── Auth ──────────────────────────────────────────────────────────────────────

service_key = getpass.getpass("Enter Supabase service role key: ").strip()
headers = {
    "apikey":        service_key,
    "Authorization": f"Bearer {service_key}",
}

# ── Data fetch ────────────────────────────────────────────────────────────────

def fetch(table, select):
    r = requests.get(
        f"{SUPABASE_URL}/rest/v1/{table}?select={select}",
        headers=headers)
    if r.status_code != 200:
        print(f"  WARNING: could not fetch {table}: {r.text[:80]}")
        return []
    return r.json()

print("\nFetching data from Supabase...")
profiles  = fetch("profiles",           "id,username,display_name,xp,level,streak_days")
feedback  = fetch("feedback_responses", "id,user_id,sus_score,open_feedback,submitted_at")
imi_raw   = fetch("imi_responses",      "id,user_id,interest_enjoyment,perceived_competence,perceived_choice,relatedness,submitted_at")
quiz_raw  = fetch("quiz_attempts",      "id,user_id,lesson_id,score,max_score,completed_at")
assessment_raw = fetch("knowledge_assessments", "id,user_id,assessment_type,score,max_score,submitted_at")
skills    = fetch("skill_mastery",      "user_id,skill_name,mastery_score,updated_at")
lessons   = fetch("lessons",            "id,title")

df_profiles = pd.DataFrame(profiles).rename(columns={"id": "user_id"}) if profiles else pd.DataFrame()
df_feedback = pd.DataFrame(feedback) if feedback else pd.DataFrame()
df_imi      = pd.DataFrame(imi_raw)  if imi_raw  else pd.DataFrame()
df_quiz     = pd.DataFrame(quiz_raw) if quiz_raw  else pd.DataFrame()
df_assessments = pd.DataFrame(assessment_raw) if assessment_raw else pd.DataFrame()
df_skills   = pd.DataFrame(skills)   if skills    else pd.DataFrame()
df_lessons  = pd.DataFrame(lessons)  if lessons   else pd.DataFrame()

# Manual thematic analysis requires human interpretation. Export a
# pseudonymised coding sheet rather than claiming automatic theme generation.
df_qualitative = pd.DataFrame(
    columns=["participant_code", "feedback", "initial_code", "theme", "researcher_notes"]
)
if not df_feedback.empty and "open_feedback" in df_feedback.columns:
    open_rows = df_feedback[
        df_feedback["open_feedback"].notna()
        & ~df_feedback["open_feedback"].str.startswith(
            ("pre_test_score:", "post_test_score:"), na=False
        )
    ][["user_id", "open_feedback"]].copy()
    participant_ids = {
        user_id: f"P{index + 1:03d}"
        for index, user_id in enumerate(sorted(df_profiles["user_id"].dropna().unique()))
    } if not df_profiles.empty else {}
    if not open_rows.empty:
        df_qualitative = pd.DataFrame({
            "participant_code": open_rows["user_id"].map(participant_ids),
            "feedback": open_rows["open_feedback"],
            "initial_code": "",
            "theme": "",
            "researcher_notes": "",
        })

print(f"  Participants:  {len(df_profiles)}")
print(f"  Feedback rows: {len(df_feedback)}")
print(f"  IMI responses: {len(df_imi)}")
print(f"  Quiz attempts: {len(df_quiz)}")
print(f"  Skill records: {len(df_skills)}")

# ── Parse pre/post tests ──────────────────────────────────────────────────────

if not df_assessments.empty:
    df_assessments["test_score"] = (
        df_assessments["score"] / df_assessments["max_score"]
    )
    df_pre = df_assessments[df_assessments["assessment_type"] == "pre"][[
        "user_id", "test_score"
    ]].rename(columns={"test_score": "pre_score"})
    df_post = df_assessments[df_assessments["assessment_type"] == "post"][[
        "user_id", "test_score"
    ]].rename(columns={"test_score": "post_score"})
else:
    df_pre = pd.DataFrame(columns=["user_id", "pre_score"])
    df_post = pd.DataFrame(columns=["user_id", "post_score"])

def classify(row):
    fb = row.get("open_feedback", "")
    if isinstance(fb, str) and fb.startswith("pre_test_score:"):  return "pre_test"
    if isinstance(fb, str) and fb.startswith("post_test_score:"): return "post_test"
    if pd.notna(row.get("sus_score")):                            return "sus_survey"
    return "other"

if not df_feedback.empty:
    df_feedback["test_type"]  = df_feedback.apply(classify, axis=1)
    df_feedback["test_score"] = df_feedback["open_feedback"].apply(
        lambda x: float(x.split(":")[1]) if isinstance(x, str) and ":" in x else None)
    if df_assessments.empty:
        df_pre  = df_feedback[df_feedback["test_type"] == "pre_test" ][["user_id","test_score"]].rename(columns={"test_score":"pre_score"})
        df_post = df_feedback[df_feedback["test_type"] == "post_test"][["user_id","test_score"]].rename(columns={"test_score":"post_score"})
    df_sus  = df_feedback[df_feedback["test_type"] == "sus_survey"][["user_id","sus_score"]]
else:
    df_sus  = pd.DataFrame(columns=["user_id","sus_score"])

# ── Build per-participant master table ────────────────────────────────────────

master = df_profiles.copy()

if not df_pre.empty:
    master = master.merge(df_pre,  on="user_id", how="left")
if not df_post.empty:
    master = master.merge(df_post, on="user_id", how="left")
if not df_sus.empty:
    master = master.merge(df_sus,  on="user_id", how="left")

if not df_imi.empty:
    imi_agg = df_imi.groupby("user_id").agg(
        imi_interest   =("interest_enjoyment",  "mean"),
        imi_competence =("perceived_competence", "mean"),
        imi_choice     =("perceived_choice",     "mean"),
        imi_relatedness=("relatedness",          "mean"),
    ).reset_index()
    imi_agg["imi_overall"] = imi_agg[["imi_interest","imi_competence","imi_choice","imi_relatedness"]].mean(axis=1)
    master = master.merge(imi_agg, on="user_id", how="left")

if not df_quiz.empty:
    df_quiz["score_ratio"] = df_quiz.apply(
        lambda row: row["score"] / row["max_score"]
        if pd.notna(row.get("max_score")) and row["max_score"] > 0 else None,
        axis=1,
    )
    quiz_agg = df_quiz.groupby("user_id").agg(
        quizzes_done  =("id",    "count"),
        avg_quiz_score=("score_ratio", "mean"),
    ).reset_index()
    master = master.merge(quiz_agg, on="user_id", how="left")

if "pre_score" in master.columns and "post_score" in master.columns:
    master["learning_gain"] = master["post_score"] - master["pre_score"]

print(f"\nParticipants with pre-test:  {master['pre_score'].notna().sum() if 'pre_score' in master.columns else 0}")
print(f"Participants with post-test: {master['post_score'].notna().sum() if 'post_score' in master.columns else 0}")
print(f"Participants with IMI:       {master['imi_overall'].notna().sum() if 'imi_overall' in master.columns else 0}")
print(f"Participants with SUS:       {master['sus_score'].notna().sum() if 'sus_score' in master.columns else 0}")

# ─────────────────────────────────────────────────────────────────────────────
# STATISTICAL ANALYSIS
# ─────────────────────────────────────────────────────────────────────────────

results = {}   # store all results for Excel export
sep = "=" * 65

def sig_label(p):
    if p < 0.001: return "*** (p<.001)"
    if p < 0.01:  return "**  (p<.01)"
    if p < 0.05:  return "*   (p<.05)"
    return "ns  (p≥.05)"

def cohens_d(a, b):
    """Paired Cohen's d."""
    diff = np.array(a) - np.array(b)
    return np.mean(diff) / np.std(diff, ddof=1) if np.std(diff, ddof=1) > 0 else 0

def effect_label(d):
    d = abs(d)
    if d >= 0.8: return "Large"
    if d >= 0.5: return "Medium"
    if d >= 0.2: return "Small"
    return "Negligible"

# ── H2: Learning Gain (pre → post test) ──────────────────────────────────────

print(f"\n{sep}")
print("H2 — LEARNING GAIN  (Pre-test → Post-test)")
print(sep)

if "pre_score" in master.columns and "post_score" in master.columns:
    paired = master.dropna(subset=["pre_score","post_score"])
    n = len(paired)
    print(f"Paired participants (completed both tests): {n}")

    if n >= 3:
        pre  = paired["pre_score"].values  * 100   # convert to %
        post = paired["post_score"].values * 100
        gain = post - pre

        # Paired t-test as specified in the interim report
        stat, p = stats.ttest_rel(pre, post)
        test_name = "Paired t-test"
        d = cohens_d(post, pre)

        print(f"\n  Pre-test  — Mean: {np.mean(pre):.1f}%  SD: {np.std(pre):.1f}%  Median: {np.median(pre):.1f}%")
        print(f"  Post-test — Mean: {np.mean(post):.1f}%  SD: {np.std(post):.1f}%  Median: {np.median(post):.1f}%")
        print(f"  Learning gain — Mean: {np.mean(gain):+.1f}%  SD: {np.std(gain):.1f}%")
        print(f"\n  {test_name}: t={stat:.3f}, p={p:.4f}  {sig_label(p)}")
        print(f"  Cohen's d = {d:.3f} → {effect_label(d)} effect")
        print(f"\n  {'✅ H2 SUPPORTED' if p < ALPHA and np.mean(gain) > 0 else '❌ H2 NOT SUPPORTED'}")

        results["H2"] = {
            "n": n, "pre_mean": np.mean(pre), "pre_sd": np.std(pre),
            "post_mean": np.mean(post), "post_sd": np.std(post),
            "mean_gain": np.mean(gain), "test": test_name,
            "t_stat": stat, "p_value": p, "cohens_d": d,
            "effect": effect_label(d), "supported": p < ALPHA and np.mean(gain) > 0,
            "pre_scores": pre.tolist(), "post_scores": post.tolist(),
            "gains": gain.tolist(), "usernames": paired["username"].tolist(),
        }
    else:
        print(f"  Insufficient paired data (n={n}). Need at least 3.")
        results["H2"] = {"n": n, "note": "Insufficient data"}
else:
    print("  No pre/post test data available yet.")
    results["H2"] = {"note": "No data"}

# ── H1: IMI Motivation vs Published Norms ────────────────────────────────────

print(f"\n{sep}")
print("H1 — IMI MOTIVATION  (vs published norm of 4.0/7)")
print(sep)

imi_subscales = ["imi_interest","imi_competence","imi_choice","imi_relatedness","imi_overall"]
imi_labels    = ["Interest/Enjoyment","Perceived Competence","Perceived Choice","Relatedness","Overall IMI"]

if "imi_overall" in master.columns and master["imi_overall"].notna().sum() >= 3:
    imi_data = master.dropna(subset=["imi_overall"])
    n_imi = len(imi_data)
    print(f"IMI respondents: {n_imi}")
    print()

    h1_results = []
    for col, label in zip(imi_subscales, imi_labels):
        if col not in imi_data.columns: continue
        vals = imi_data[col].dropna().values
        if len(vals) < 3: continue

        # Descriptive statistics comparison vs published norm (interim report method)
        mean_val = np.mean(vals)
        sd_val   = np.std(vals, ddof=1)
        se_val   = sd_val / np.sqrt(len(vals))
        ci_lo    = mean_val - 1.96 * se_val
        ci_hi    = mean_val + 1.96 * se_val
        above_norm = mean_val > IMI_NORM
        d_one    = (mean_val - IMI_NORM) / sd_val if sd_val > 0 else 0

        print(f"  {label:<24} Mean={mean_val:.2f}  SD={sd_val:.2f}  "
              f"95% CI=[{ci_lo:.2f},{ci_hi:.2f}]  norm={IMI_NORM}  "
              f"{'ABOVE' if above_norm else 'BELOW'} norm  d={d_one:.2f}")

        h1_results.append({
            "subscale": label, "mean": mean_val, "sd": sd_val,
            "n": len(vals), "ci_95_lo": ci_lo, "ci_95_hi": ci_hi,
            "vs_norm": IMI_NORM, "above_norm": above_norm,
            "cohens_d": d_one, "effect": effect_label(d_one),
        })

    overall_vals = imi_data["imi_overall"].dropna().values
    supported = np.mean(overall_vals) > IMI_NORM
    print(f"\n  {'✅ H1 SUPPORTED' if supported else '❌ H1 NOT SUPPORTED'} "
          f"(overall mean {np.mean(overall_vals):.2f} {'>' if supported else '<='} norm {IMI_NORM})")
    results["H1"] = h1_results
else:
    print("  No IMI data available yet.")
    results["H1"] = []

# ── H3: SDT Scores vs Engagement (Spearman Correlation) ──────────────────────

print(f"\n{sep}")
print("H3 — SDT vs ENGAGEMENT  (Spearman correlations, supplementary quantitative analysis)")
print(sep)

engagement_vars = {
    "xp":            "XP Earned",
    "quizzes_done":  "Quizzes Completed",
    "avg_quiz_score":"Avg Quiz Score",
    "streak_days":   "Streak Days",
}
imi_vars = {
    "imi_interest":    "Interest/Enjoyment",
    "imi_competence":  "Perceived Competence",
    "imi_choice":      "Perceived Choice",
    "imi_relatedness": "Relatedness",
}

h3_results = []
corr_matrix_data = {}

if "imi_overall" in master.columns and master["imi_overall"].notna().sum() >= 3:
    print()
    for imi_col, imi_label in imi_vars.items():
        if imi_col not in master.columns: continue
        for eng_col, eng_label in engagement_vars.items():
            if eng_col not in master.columns: continue
            sub = master.dropna(subset=[imi_col, eng_col])
            if len(sub) < 3: continue
            r, p = stats.spearmanr(sub[imi_col], sub[eng_col])
            sig = "*" if p < ALPHA else ""
            print(f"  {imi_label:<24} ↔ {eng_label:<22} r={r:+.3f}  p={p:.4f}  {sig_label(p)}")
            h3_results.append({
                "imi_subscale": imi_label, "engagement_var": eng_label,
                "n": len(sub), "spearman_r": r, "p_value": p,
                "significant": p < ALPHA,
            })
            corr_matrix_data[f"{imi_label[:8]}↔{eng_label[:8]}"] = r

    sig_count = sum(1 for r in h3_results if r["significant"])
    print(f"\n  Significant correlations: {sig_count}/{len(h3_results)}")
    print(f"  {'✅ H3 SUPPORTED' if sig_count > 0 else '❌ H3 NOT SUPPORTED'}")
    results["H3"] = h3_results
else:
    print("  Insufficient IMI data for correlation analysis.")
    results["H3"] = []

# ── SUS Usability Analysis ────────────────────────────────────────────────────

print(f"\n{sep}")
print("SUS USABILITY ANALYSIS  (benchmark = 68)")
print(sep)

if "sus_score" in master.columns and master["sus_score"].notna().sum() >= 1:
    sus_vals = master["sus_score"].dropna().values
    n_sus = len(sus_vals)
    mean_sus = np.mean(sus_vals)

    def sus_grade(s):
        if s >= 85: return "Excellent"
        if s >= 72: return "Good"
        if s >= 52: return "OK"
        if s >= 38: return "Poor"
        return "Awful"

    print(f"\n  N={n_sus}  Mean={mean_sus:.1f}  SD={np.std(sus_vals):.1f}  "
          f"Median={np.median(sus_vals):.1f}")
    print(f"  Grade: {sus_grade(mean_sus)}  ({'above' if mean_sus >= SUS_BENCH else 'below'} benchmark of {SUS_BENCH})")

    if n_sus >= 3:
        stat, p = stats.wilcoxon(sus_vals - SUS_BENCH)
        print(f"  Wilcoxon vs benchmark: stat={stat:.3f}, p={p:.4f}  {sig_label(p)}")

    results["SUS"] = {
        "n": n_sus, "mean": mean_sus, "sd": np.std(sus_vals),
        "median": np.median(sus_vals), "grade": sus_grade(mean_sus),
        "above_benchmark": mean_sus >= SUS_BENCH, "scores": sus_vals.tolist(),
    }
else:
    print("  No SUS data available yet.")
    results["SUS"] = {}

# ── Skill mastery summary ─────────────────────────────────────────────────────

print(f"\n{sep}")
print("SKILL MASTERY SUMMARY")
print(sep)

if not df_skills.empty:
    skill_summary = df_skills.groupby("skill_name")["mastery_score"].agg(
        ["mean","std","min","max","count"]).round(3).reset_index()
    skill_summary.columns = ["Skill","Mean","SD","Min","Max","N"]
    skill_summary = skill_summary.sort_values("Mean", ascending=False)
    print(f"\n  {'Skill':<22} {'Mean':>6} {'SD':>6} {'Min':>6} {'Max':>6} {'N':>4}")
    print("  " + "-"*52)
    for _, row in skill_summary.iterrows():
        bar = "█" * int(row["Mean"] * 20)
        print(f"  {row['Skill']:<22} {row['Mean']:>6.3f} {row['SD']:>6.3f} "
              f"{row['Min']:>6.3f} {row['Max']:>6.3f} {int(row['N']):>4}  {bar}")
    results["skills"] = skill_summary

# ─────────────────────────────────────────────────────────────────────────────
# FIGURES
# ─────────────────────────────────────────────────────────────────────────────

print(f"\n{sep}")
print("Generating charts...")

ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
pdf_path  = f"C:/Users/munug/arete_app/arete_analysis_charts_{ts}.pdf"
xlsx_path = f"C:/Users/munug/arete_app/arete_analysis_results_{ts}.xlsx"

NAVY   = "#1A1A6E"
TEAL   = "#00D4AA"
PURPLE = "#6C3DE0"
GOLD   = "#FFD700"
RED    = "#E55934"
GREY   = "#CCCCCC"

plt.rcParams.update({
    "font.family": "DejaVu Sans", "font.size": 10,
    "axes.spines.top": False, "axes.spines.right": False,
    "figure.facecolor": "white", "axes.facecolor": "#F8F9FA",
    "axes.grid": True, "grid.color": "#E0E0E0", "grid.linewidth": 0.6,
})

with PdfPages(pdf_path) as pdf:

    # ── Figure 1: Pre vs Post Test Scores ────────────────────────────────────
    if "H2" in results and "pre_scores" in results["H2"]:
        h2 = results["H2"]
        fig, axes = plt.subplots(1, 2, figsize=(12, 5))
        fig.suptitle("H2 — Knowledge Assessment: Pre-Test vs Post-Test",
                     fontsize=14, fontweight="bold", color=NAVY, y=1.01)

        # Paired lines chart
        ax = axes[0]
        usernames = h2["usernames"]
        for i, (pre, post, name) in enumerate(zip(h2["pre_scores"], h2["post_scores"], usernames)):
            color = TEAL if post > pre else RED
            ax.plot([0, 1], [pre, post], "o-", color=color, alpha=0.7,
                    linewidth=2, markersize=8, label=name if i < 8 else "")
        ax.set_xticks([0, 1])
        ax.set_xticklabels(["Pre-Test", "Post-Test"], fontsize=11)
        ax.set_ylabel("Score (%)")
        ax.set_ylim(0, 105)
        ax.set_title("Individual Score Trajectories\n(green = improvement, red = decline)")
        ax.axhline(y=h2["pre_mean"],  color=GREY, linestyle="--", linewidth=1, alpha=0.7)
        ax.axhline(y=h2["post_mean"], color=TEAL, linestyle="--", linewidth=1, alpha=0.7)

        # Bar chart with error bars
        ax2 = axes[1]
        means = [h2["pre_mean"], h2["post_mean"]]
        sds   = [h2["pre_sd"],   h2["post_sd"]]
        bars  = ax2.bar(["Pre-Test", "Post-Test"], means, color=[PURPLE, TEAL],
                        alpha=0.85, width=0.5, zorder=3)
        ax2.errorbar(["Pre-Test", "Post-Test"], means, yerr=sds,
                     fmt="none", color=NAVY, capsize=6, linewidth=2, zorder=4)
        for bar, mean in zip(bars, means):
            ax2.text(bar.get_x() + bar.get_width()/2, mean + 2,
                     f"{mean:.1f}%", ha="center", va="bottom",
                     fontweight="bold", fontsize=11)
        ax2.set_ylabel("Mean Score (%)")
        ax2.set_ylim(0, 105)
        ax2.set_title(f"Group Means (n={h2['n']})\n"
                      f"Mean gain: {h2['mean_gain']:+.1f}%  |  "
                      f"p={h2['p_value']:.4f}  |  d={h2['cohens_d']:.2f} ({h2['effect']})")
        ax2.axhline(y=50, color=GREY, linestyle=":", linewidth=1, label="Chance (50%)")
        ax2.legend(fontsize=9)

        plt.tight_layout()
        pdf.savefig(fig, bbox_inches="tight")
        plt.close()

    # ── Figure 2: IMI Subscales ───────────────────────────────────────────────
    if "H1" in results and results["H1"]:
        h1 = results["H1"]
        fig, axes = plt.subplots(1, 2, figsize=(12, 5))
        fig.suptitle("H1 — IMI Motivation Scores vs Published Norm (4.0/7)",
                     fontsize=14, fontweight="bold", color=NAVY)

        labels = [r["subscale"] for r in h1]
        means  = [r["mean"]     for r in h1]
        sds    = [r["sd"]       for r in h1]
        colors = [TEAL if m > IMI_NORM else RED for m in means]

        ax = axes[0]
        bars = ax.barh(labels, means, color=colors, alpha=0.85, height=0.5)
        ax.axvline(x=IMI_NORM, color=NAVY, linestyle="--",
                   linewidth=2, label=f"Norm = {IMI_NORM}")
        ax.errorbar(means, labels, xerr=sds, fmt="none",
                    color=NAVY, capsize=4, linewidth=1.5)
        for bar, m in zip(bars, means):
            ax.text(m + 0.05, bar.get_y() + bar.get_height()/2,
                    f"{m:.2f}", va="center", fontsize=9, fontweight="bold")
        ax.set_xlabel("Score (1–7)")
        ax.set_xlim(0, 7.5)
        ax.set_title("IMI Subscale Means")
        ax.legend()

        # Radar chart
        ax2 = axes[1]
        subscale_data = [r for r in h1 if r["subscale"] != "Overall IMI"]
        cats   = [r["subscale"].replace("/", "\n") for r in subscale_data]
        values = [r["mean"] for r in subscale_data]
        N      = len(cats)
        angles = [n / float(N) * 2 * np.pi for n in range(N)]
        angles += angles[:1]
        values += values[:1]
        norm_line = [IMI_NORM] * (N + 1)

        ax2 = fig.add_subplot(122, polar=True)
        ax2.plot(angles, values, "o-", linewidth=2, color=TEAL, label="Arete")
        ax2.fill(angles, values, alpha=0.2, color=TEAL)
        ax2.plot(angles, norm_line, "--", linewidth=1.5, color=GREY, label=f"Norm ({IMI_NORM})")
        ax2.set_xticks(angles[:-1])
        ax2.set_xticklabels(cats, size=9)
        ax2.set_ylim(0, 7)
        ax2.set_title("IMI Radar", pad=20)
        ax2.legend(loc="upper right", bbox_to_anchor=(1.3, 1.1))

        plt.tight_layout()
        pdf.savefig(fig, bbox_inches="tight")
        plt.close()

    # ── Figure 3: Spearman Correlation Heatmap ────────────────────────────────
    if "H3" in results and results["H3"]:
        h3 = results["H3"]
        imi_lbls = list(dict.fromkeys(r["imi_subscale"]   for r in h3))
        eng_lbls = list(dict.fromkeys(r["engagement_var"] for r in h3))
        matrix   = pd.DataFrame(index=imi_lbls, columns=eng_lbls, dtype=float)
        pmatrix  = pd.DataFrame(index=imi_lbls, columns=eng_lbls, dtype=float)

        for r in h3:
            matrix.loc[r["imi_subscale"],   r["engagement_var"]] = r["spearman_r"]
            pmatrix.loc[r["imi_subscale"],  r["engagement_var"]] = r["p_value"]

        fig, ax = plt.subplots(figsize=(10, 5))
        sns.heatmap(matrix.astype(float), annot=True, fmt=".2f", cmap="RdYlGn",
                    center=0, vmin=-1, vmax=1, ax=ax,
                    linewidths=0.5, linecolor="white",
                    cbar_kws={"label": "Spearman r"})

        # Asterisks for significance
        for i, imi in enumerate(imi_lbls):
            for j, eng in enumerate(eng_lbls):
                p = pmatrix.loc[imi, eng]
                if pd.notna(p) and p < ALPHA:
                    ax.text(j + 0.5, i + 0.85, "*", ha="center",
                            fontsize=14, fontweight="bold", color="black")

        ax.set_title("H3 — Spearman Correlations: IMI Subscales vs Engagement\n"
                     "(* = p<.05, green = positive, red = negative)",
                     fontsize=12, fontweight="bold", color=NAVY)
        ax.set_xlabel("Engagement Metrics")
        ax.set_ylabel("IMI Subscales")
        plt.tight_layout()
        pdf.savefig(fig, bbox_inches="tight")
        plt.close()

    # ── Figure 4: SUS Score ───────────────────────────────────────────────────
    if "SUS" in results and results["SUS"] and "scores" in results["SUS"]:
        sus = results["SUS"]
        fig, axes = plt.subplots(1, 2, figsize=(12, 5))
        fig.suptitle("SUS Usability Analysis", fontsize=14,
                     fontweight="bold", color=NAVY)

        # Individual scores
        ax = axes[0]
        scores = sorted(sus["scores"])
        colors = [TEAL if s >= SUS_BENCH else RED for s in scores]
        bars = ax.bar(range(1, len(scores)+1), scores, color=colors, alpha=0.85, width=0.5)
        ax.axhline(y=SUS_BENCH, color=NAVY, linestyle="--",
                   linewidth=2, label=f"Benchmark ({SUS_BENCH})")
        ax.axhline(y=sus["mean"], color=GOLD, linestyle="-",
                   linewidth=2, label=f"Mean ({sus['mean']:.1f})")
        for bar, s in zip(bars, scores):
            ax.text(bar.get_x() + bar.get_width()/2, s + 1,
                    f"{s:.0f}", ha="center", fontsize=9, fontweight="bold")
        ax.set_xlabel("Participant")
        ax.set_ylabel("SUS Score (0–100)")
        ax.set_ylim(0, 110)
        ax.set_title(f"Individual SUS Scores (n={sus['n']})")
        ax.legend()

        # Grade gauge
        ax2 = axes[1]
        grade_bands = [
            (0,  38,  "#EF5350", "Awful"),
            (38, 52,  "#FF7043", "Poor"),
            (52, 68,  "#FFA726", "OK"),
            (68, 72,  "#FFEE58", "Good"),
            (72, 85,  "#66BB6A", "Good"),
            (85, 100, "#26A69A", "Excellent"),
        ]
        for lo, hi, color, label in grade_bands:
            ax2.barh(0, hi-lo, left=lo, color=color, height=0.4, alpha=0.7)
        ax2.axvline(x=sus["mean"], color=NAVY, linewidth=4, zorder=5)
        ax2.text(sus["mean"], 0.25, f"{sus['mean']:.1f}\n{sus['grade']}",
                 ha="center", fontsize=13, fontweight="bold", color=NAVY)
        ax2.set_xlim(0, 100)
        ax2.set_yticks([])
        ax2.set_xlabel("SUS Score")
        ax2.set_title("Usability Grade")
        for lo, hi, _, label in grade_bands:
            ax2.text((lo+hi)/2, -0.3, label, ha="center", fontsize=8, color="#333")

        plt.tight_layout()
        pdf.savefig(fig, bbox_inches="tight")
        plt.close()

    # ── Figure 5: Skill Mastery ───────────────────────────────────────────────
    if "skills" in results and not results["skills"].empty:
        skill_df = results["skills"]
        fig, ax = plt.subplots(figsize=(10, 5))
        bars = ax.barh(skill_df["Skill"], skill_df["Mean"],
                       color=TEAL, alpha=0.85, height=0.5)
        ax.errorbar(skill_df["Mean"], skill_df["Skill"],
                    xerr=skill_df["SD"], fmt="none",
                    color=NAVY, capsize=4, linewidth=1.5)
        for bar, mean in zip(bars, skill_df["Mean"]):
            ax.text(mean + 0.01, bar.get_y() + bar.get_height()/2,
                    f"{mean:.3f}", va="center", fontsize=9)
        ax.set_xlabel("Mean Mastery Score (0–1)")
        ax.set_xlim(0, 1.1)
        ax.axvline(x=0.5, color=GREY, linestyle="--", linewidth=1, label="0.5 threshold")
        ax.set_title("Skill Mastery — Group Means with SD",
                     fontsize=12, fontweight="bold", color=NAVY)
        ax.legend()
        plt.tight_layout()
        pdf.savefig(fig, bbox_inches="tight")
        plt.close()

    # ── Cover page ────────────────────────────────────────────────────────────
    fig = plt.figure(figsize=(12, 5))
    fig.text(0.5, 0.75, "Arete Research Study — Statistical Analysis",
             ha="center", fontsize=16, fontweight="bold", color=NAVY)
    fig.text(0.5, 0.60, "Newcastle University · MSc Computer Science · CSC8639",
             ha="center", fontsize=12, color="#555")
    fig.text(0.5, 0.50, f"Researcher: Shalem Raj Munugala",
             ha="center", fontsize=11, color="#555")
    fig.text(0.5, 0.42, "Supervisor: Dr Sara Johansson Fernstad",
             ha="center", fontsize=11, color="#555")
    fig.text(0.5, 0.28, f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}",
             ha="center", fontsize=10, color="#888")
    pdf.savefig(fig, bbox_inches="tight")
    plt.close()

print(f"Charts saved: {pdf_path}")

# ─────────────────────────────────────────────────────────────────────────────
# EXCEL RESULTS EXPORT
# ─────────────────────────────────────────────────────────────────────────────

print("Writing results to Excel...")

header_fill = PatternFill("solid", fgColor="2E4057")
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
title_fill  = PatternFill("solid", fgColor="1A1A6E")
title_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin        = Side(style="thin", color="CCCCCC")
border      = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_sheet(ws, title):
    last_col = get_column_letter(ws.max_column)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font      = title_font
    ws["A1"].fill      = title_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28
    for cell in ws[2]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[2].height = 32
    alt = PatternFill("solid", fgColor="F5F7FA")
    for row_idx, row in enumerate(ws.iter_rows(min_row=3), start=1):
        for cell in row:
            cell.alignment = center
            cell.border    = border
            if row_idx % 2 == 0:
                cell.fill = alt
    for col in ws.columns:
        first = col[0]
        if not hasattr(first, "column_letter"):
            continue
        max_len = max((len(str(c.value)) for c in col if c.value and hasattr(c, "column_letter")), default=10)
        ws.column_dimensions[first.column_letter].width = min(max_len + 4, 36)

with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:

    # Sheet 1 — H2 Learning Gain
    if "H2" in results and "pre_scores" in results["H2"]:
        h2 = results["H2"]
        df_h2 = pd.DataFrame({
            "Participant": h2["usernames"],
            "Pre-Test (%)": [f"{s:.1f}" for s in h2["pre_scores"]],
            "Post-Test (%)": [f"{s:.1f}" for s in h2["post_scores"]],
            "Learning Gain (%)": [f"{g:+.1f}" for g in h2["gains"]],
        })
        summary_row = pd.DataFrame([{
            "Participant": "GROUP MEAN",
            "Pre-Test (%)": f"{h2['pre_mean']:.1f}",
            "Post-Test (%)": f"{h2['post_mean']:.1f}",
            "Learning Gain (%)": f"{h2['mean_gain']:+.1f}",
        }])
        df_h2 = pd.concat([df_h2, summary_row], ignore_index=True)
        df_h2.to_excel(writer, sheet_name="H2 Learning Gain", index=False, startrow=1)
        ws = writer.sheets["H2 Learning Gain"]
        style_sheet(ws, f"H2 — Learning Gain | {h2['test']} p={h2['p_value']:.4f} | d={h2['cohens_d']:.2f} ({h2['effect']}) | n={h2['n']}")

    # Sheet 2 — H1 IMI
    if "H1" in results and results["H1"]:
        pd.DataFrame(results["H1"]).to_excel(writer, sheet_name="H1 IMI Scores", index=False, startrow=1)
        style_sheet(writer.sheets["H1 IMI Scores"], "H1 — IMI Motivation Scores vs Published Norm (4.0/7)")

    # Sheet 3 — H3 Correlations
    if "H3" in results and results["H3"]:
        pd.DataFrame(results["H3"]).to_excel(writer, sheet_name="H3 Correlations", index=False, startrow=1)
        style_sheet(writer.sheets["H3 Correlations"], "H3 — Spearman Correlations: SDT vs Engagement")

    # Sheet 4 — SUS
    if "SUS" in results and results["SUS"] and "scores" in results["SUS"]:
        sus = results["SUS"]
        df_sus_out = pd.DataFrame({
            "Metric": ["N","Mean","SD","Median","Grade","Above Benchmark (68)"],
            "Value":  [sus["n"], f"{sus['mean']:.1f}", f"{sus['sd']:.1f}",
                       f"{sus['median']:.1f}", sus["grade"],
                       "Yes" if sus["above_benchmark"] else "No"],
        })
        df_sus_out.to_excel(writer, sheet_name="SUS Usability", index=False, startrow=1)
        style_sheet(writer.sheets["SUS Usability"], "SUS Usability Analysis (Benchmark = 68)")

    # Sheet 5 — Skill Mastery
    if "skills" in results and not results["skills"].empty:
        results["skills"].to_excel(writer, sheet_name="Skill Mastery", index=False, startrow=1)
        style_sheet(writer.sheets["Skill Mastery"], "Skill Mastery — Group Summary")

    # Human coding remains required; these columns make that process explicit
    # and auditable without exposing account identifiers.
    if not df_qualitative.empty:
        df_qualitative.to_excel(
            writer, sheet_name="Qualitative Coding", index=False, startrow=1
        )
        style_sheet(
            writer.sheets["Qualitative Coding"],
            "Manual Thematic Analysis Coding Sheet",
        )

print(f"Results saved: {xlsx_path}")

print(f"\n{sep}")
print("ANALYSIS COMPLETE")
print(sep)
print(f"  Charts PDF : {pdf_path}")
print(f"  Results XLS: {xlsx_path}")
print()
print("  Include in your dissertation:")
print("  • H2 table from Sheet 'H2 Learning Gain'")
print("  • Figure 1 (pre/post chart) from PDF page 1")
print("  • H1 table from Sheet 'H1 IMI Scores'")
print("  • Figure 3 (correlation heatmap) from PDF page 3")
print("  • SUS grade from Sheet 'SUS Usability'")
